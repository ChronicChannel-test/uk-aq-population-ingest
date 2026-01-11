from dataclasses import dataclass
from datetime import date
import csv
import io
import json
from pathlib import Path
import re
from typing import Iterable

import requests

from nomis_api.geo import normalize_geo_type
from nomis_api.ingest import parse_reference_date


@dataclass(frozen=True)
class CatalogueEntry:
    geography_level: str
    geography_vintage: str
    coverage: str
    nomis_dataset_api_ref: str | None
    nomis_dataset_keyfamily_id: str | None
    nomis_geography_type_code: str | None
    typical_update_cycle: str | None
    latest_reference_period_on_nomis: str | None
    next_release_note: str | None
    notes: str | None
    primary_source_org: str | None
    primary_source_dataset_page: str | None
    primary_source_download_csv: str | None
    primary_source_download_xlsx: str | None
    primary_source_api_example: str | None
    primary_source_years_available: str | None
    primary_source_update_frequency: str | None
    primary_source_notes: str | None


@dataclass(frozen=True)
class IngestOverride:
    dataset_id: str
    source_url: str | None = None
    format: str | None = None
    geo_code_column: str | None = None
    time_column: str | None = None
    value_column: str | None = None
    measure_column: str | None = None
    jsonstat_geo_dimension: str | None = None
    jsonstat_time_dimension: str | None = None
    jsonstat_value_dimension: str | None = None


def load_catalogue(path: Path) -> list[CatalogueEntry]:
    rows = []
    with path.open(newline="", encoding="utf-8") as handle:
        reader = csv.DictReader(handle)
        for row in reader:
            rows.append(
                CatalogueEntry(
                    geography_level=row["geography_level"],
                    geography_vintage=row["geography_vintage"],
                    coverage=row["coverage"],
                    nomis_dataset_api_ref=row.get("nomis_dataset_api_ref") or None,
                    nomis_dataset_keyfamily_id=row.get("nomis_dataset_keyfamily_id") or None,
                    nomis_geography_type_code=row.get("nomis_geography_type_code") or None,
                    typical_update_cycle=row.get("typical_update_cycle") or None,
                    latest_reference_period_on_nomis=row.get("latest_reference_period_on_nomis") or None,
                    next_release_note=row.get("next_release_note") or None,
                    notes=row.get("notes") or None,
                    primary_source_org=row.get("primary_source_org") or None,
                    primary_source_dataset_page=row.get("primary_source_dataset_page") or None,
                    primary_source_download_csv=row.get("primary_source_download_csv") or None,
                    primary_source_download_xlsx=row.get("primary_source_download_xlsx") or None,
                    primary_source_api_example=row.get("primary_source_api_example") or None,
                    primary_source_years_available=row.get("primary_source_years_available") or None,
                    primary_source_update_frequency=row.get("primary_source_update_frequency") or None,
                    primary_source_notes=row.get("primary_source_notes") or None,
                )
            )
    return rows


def load_ingest_overrides(path: Path) -> dict[str, IngestOverride]:
    payload = json.loads(path.read_text())
    overrides = {}
    for item in payload.get("overrides", []):
        override = IngestOverride(**item)
        overrides[override.dataset_id] = override
    return overrides


def assign_source_prefix(entry: CatalogueEntry) -> str:
    if entry.nomis_dataset_api_ref or entry.nomis_dataset_keyfamily_id or entry.nomis_geography_type_code:
        return "nomis"
    coverage = entry.coverage.lower()
    if "scotland" in coverage:
        return "nrs"
    if "northern ireland" in coverage:
        return "nisra"
    if "england" in coverage or "wales" in coverage:
        return "nomis"
    return "nomis"


def slugify(value: str) -> str:
    value = value.strip().lower()
    value = re.sub(r"[^a-z0-9]+", "_", value)
    return value.strip("_")


def build_dataset_id(prefix: str, entry: CatalogueEntry) -> str:
    parts = [
        prefix,
        slugify(entry.geography_level),
        slugify(entry.geography_vintage),
        slugify(entry.coverage),
    ]
    return "_".join([part for part in parts if part])


def pick_source_url(entry: CatalogueEntry) -> str | None:
    return (
        entry.primary_source_download_csv
        or entry.primary_source_download_xlsx
        or entry.primary_source_api_example
    )


def infer_format(url: str) -> str:
    lowered = url.lower()
    if lowered.endswith(".csv"):
        return "csv"
    if lowered.endswith(".xlsx") or lowered.endswith(".xls"):
        return "xlsx"
    if ".json" in lowered:
        return "jsonstat"
    return "csv"


def fetch_csv_rows(url: str) -> list[dict]:
    response = requests.get(url, timeout=120)
    response.raise_for_status()
    text = response.text
    reader = csv.DictReader(io.StringIO(text))
    return list(reader)


def fetch_xlsx_rows(url: str) -> list[dict]:
    from openpyxl import load_workbook

    response = requests.get(url, timeout=120)
    response.raise_for_status()
    workbook = load_workbook(io.BytesIO(response.content), read_only=True, data_only=True)
    sheet = workbook.active
    rows = list(sheet.iter_rows(values_only=True))
    if not rows:
        return []
    headers = [str(value).strip() if value is not None else "" for value in rows[0]]
    data_rows = []
    for row in rows[1:]:
        mapped = {}
        for idx, header in enumerate(headers):
            if header:
                mapped[header] = row[idx] if idx < len(row) else None
        data_rows.append(mapped)
    return data_rows


def ordered_categories(category_index: dict) -> list[str]:
    return [key for key, _ in sorted(category_index.items(), key=lambda item: item[1])]


def fetch_jsonstat_rows(
    url: str,
    geo_dimension: str,
    time_dimension: str,
    value_dimension: str | None,
) -> list[dict]:
    response = requests.get(url, timeout=120)
    response.raise_for_status()
    payload = response.json()
    dataset = payload.get("dataset", payload)
    dimension = dataset["dimension"]
    ids = dataset["id"]
    sizes = dataset["size"]
    values = dataset["value"]

    id_to_pos = {dimension_id: idx for idx, dimension_id in enumerate(ids)}
    geo_pos = id_to_pos[geo_dimension]
    time_pos = id_to_pos[time_dimension]
    value_pos = id_to_pos[value_dimension] if value_dimension and value_dimension in id_to_pos else None

    geo_cats = ordered_categories(dimension[geo_dimension]["category"]["index"])
    time_cats = ordered_categories(dimension[time_dimension]["category"]["index"])
    value_cats = (
        ordered_categories(dimension[value_dimension]["category"]["index"]) if value_pos is not None else [None]
    )

    rows = []
    for geo in geo_cats:
        for time in time_cats:
            for value_key in value_cats:
                coords = [0] * len(ids)
                coords[geo_pos] = dimension[geo_dimension]["category"]["index"][geo]
                coords[time_pos] = dimension[time_dimension]["category"]["index"][time]
                if value_pos is not None:
                    coords[value_pos] = dimension[value_dimension]["category"]["index"][value_key]
                index = 0
                for size, coord in zip(sizes, coords):
                    index = index * size + coord
                value = values[index]
                row = {geo_dimension: geo, time_dimension: time, "value": value}
                if value_pos is not None and value_dimension is not None:
                    row[value_dimension] = value_key
                rows.append(row)
    return rows


def infer_column(headers: Iterable[str], candidates: Iterable[str]) -> str | None:
    lower_headers = {header.lower(): header for header in headers}
    for candidate in candidates:
        if candidate.lower() in lower_headers:
            return lower_headers[candidate.lower()]
    for header in headers:
        header_lower = header.lower()
        for candidate in candidates:
            if candidate.lower() in header_lower:
                return header
    return None


def parse_reference_date_flexible(value: str) -> date:
    try:
        return parse_reference_date(str(value))
    except ValueError:
        match = re.search(r"(19|20)\\d{2}", str(value))
        if match:
            return date(int(match.group(0)), 6, 30)
    raise ValueError(f"Unsupported time value: {value}")


def map_rows(
    rows: Iterable[dict],
    geo_type: str,
    dataset_id: str,
    geo_code_column: str,
    time_column: str,
    value_column: str,
    measure_column: str | None,
) -> list[dict]:
    mapped = []
    for row in rows:
        reference_date = parse_reference_date_flexible(row[time_column])
        mapped.append(
            {
                "geo_type": normalize_geo_type(geo_type),
                "geo_code": str(row[geo_code_column]).strip(),
                "reference_date": reference_date.isoformat(),
                "population_value": int(float(row[value_column])),
                "dataset_id": dataset_id,
                "measure": row.get(measure_column) if measure_column else None,
            }
        )
    return mapped
